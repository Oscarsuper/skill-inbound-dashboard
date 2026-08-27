# -*- coding: utf-8 -*-
"""Evalúa las fórmulas COUNTIFS del libro contra la hoja DATOS, sin Excel.
No basta con que la fórmula 'corra': tiene que dar el número correcto."""
import re, pandas as pd
from openpyxl import load_workbook

wb = load_workbook('xls/Tablero_NPS_Junio_Agosto_2026.xlsx')
d  = pd.read_pickle('xls/base.pkl')
COL = {c: i+1 for i, c in enumerate(d.columns)}
LET = {i+1: c for i, c in enumerate(d.columns)}   # 1->Fecha ... 9->Clasificacion
def col_de(letra):
    from openpyxl.utils import column_index_from_string
    return LET[column_index_from_string(letra)]

RE_CIF = re.compile(r'COUNTIFS?\((.*?)\)(?=[-)/,]|$)')
def evalua_countifs(arg):
    """arg: 'DATOS!$C$2:$C$40739,"Inbound",DATOS!$B$2:$B$40739,"Junio"'"""
    partes = [p.strip() for p in re.split(r',(?=DATOS!|")', arg)]
    m = d
    i = 0
    toks = re.findall(r'DATOS!\$([A-Z]+)\$\d+:\$[A-Z]+\$\d+|"([^"]*)"', arg)
    pares = []
    while i < len(toks):
        letra = toks[i][0]; i += 1
        crit  = toks[i][1]; i += 1
        pares.append((col_de(letra), crit))
    for c, v in pares:
        m = m[m[c].astype(str) == v]
    return len(m)

def valor(formula):
    """Resuelve =IFERROR((COUNTIFS..-COUNTIFS..)/COUNTIFS..,"") y =COUNTIFS(..)"""
    f = formula.lstrip('=')
    cifs = RE_CIF.findall(f)
    if not cifs: return None
    vals = [evalua_countifs(a) for a in cifs]
    if len(vals) == 3:
        # nps_formula escribe en este orden: promotores, detractores, total
        pro, det, tot = vals
        return (pro - det) / tot if tot else None
    if len(vals) == 1:
        return vals[0]
    return None

errores = 0; revisadas = 0
def chequear(hoja, celda, esperado, etiqueta, tol=1e-9):
    global errores, revisadas
    f = wb[hoja][celda].value
    if not isinstance(f, str) or not f.startswith('='):
        print(f"  ✗ {etiqueta}: {celda} no es fórmula ({f!r})"); errores += 1; return
    v = valor(f)
    revisadas += 1
    if v is None:
        print(f"  ? {etiqueta}: no pude evaluar {celda}"); return
    ok = abs(v - esperado) < (tol if esperado < 2 else 0.5)
    if not ok:
        print(f"  ✗ {etiqueta} ({celda}): fórmula da {v}, esperado {esperado}"); errores += 1
    else:
        print(f"  ✓ {etiqueta}: {v:.4f}" if esperado < 2 else f"  ✓ {etiqueta}: {int(v):,}")

def nps_de(sub):
    if not len(sub): return None
    return (sub['Clasificacion'].eq('Promotores').sum()
            - sub['Clasificacion'].eq('Detractores').sum()) / len(sub)

print("── TABLERO: indicadores generales ──")
chequear('TABLERO','A6', nps_de(d), 'NPS general')
chequear('TABLERO','E6', int(d['Clasificacion'].eq('Detractores').sum()), 'Detractores')
chequear('TABLERO','I6', nps_de(d[d.Skill=='Inbound']), 'NPS Inbound')
chequear('TABLERO','K6', nps_de(d[d.Skill=='Chat']), 'NPS Chat')

print("\n── TABLERO: comparativa mensual ──")
fila = 12
for sk in ['Inbound','Chat']:
    for m in ['Junio','Julio','Agosto']:
        s = d[(d.Skill==sk)&(d.Mes==m)]
        chequear('TABLERO', f'C{fila}', len(s), f'{sk} {m} encuestas')
        chequear('TABLERO', f'H{fila}', nps_de(s), f'{sk} {m} NPS')
        fila += 1
    fila += 1

print("\n── TIPIFICACIÓN: primeras filas de cada hoja ──")
for sk in ['INBOUND','CHAT']:
    hoja = f'TIPIFICACIÓN {sk}'
    nombre = wb[hoja]['A6'].value
    s = d[(d.Skill==sk.capitalize())&(d['Proceso']==nombre)]
    chequear(hoja,'B6', len(s), f'{sk} · {nombre[:22]} encuestas')
    chequear(hoja,'C6', int(s['Clasificacion'].eq('Detractores').sum()),
             f'{sk} · {nombre[:22]} detractores')
    chequear(hoja,'F6', nps_de(s), f'{sk} · {nombre[:22]} NPS')

print("\n── AGENTES: primer y último agente de cada hoja ──")
for sk in ['INBOUND','CHAT']:
    hoja = f'AGENTES {sk}'
    ult = max(r for r in range(7, wb[hoja].max_row+1) if wb[hoja][f'B{r}'].value)
    for r in (7, ult):
        a = wb[hoja][f'B{r}'].value
        s = d[(d.Skill==sk.capitalize())&(d.Agente==a)]
        chequear(hoja, f'C{r}', len(s), f'{sk} · {a[:26]} encuestas')
        chequear(hoja, f'G{r}', nps_de(s), f'{sk} · {a[:26]} NPS')

print(f"\n{revisadas} fórmulas evaluadas · {errores} con diferencia")
