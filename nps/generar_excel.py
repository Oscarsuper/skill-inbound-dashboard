# -*- coding: utf-8 -*-
"""Tablero de NPS en Excel: los datos en una hoja, el resto son tableros
que se calculan con fórmulas sobre esa hoja."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.comments import Comment

AZUL   = '2B0B6B'; AZUL2 = '4A1D96'; AZUL3 = '7C3AED'
LILA   = 'A78BFA'; LILA_LT='EDE9FE'; BG    = 'F8F5FF'
LN     = 'DDD6FE'; TXT   = '6B7280'
OK     = '16A34A'; BAD   = 'DC2626'; WN    = 'D97706'
FUENTE = 'Arial'

base = pd.read_pickle('xls/base.pkl')
MESES  = ['Junio','Julio','Agosto']
SKILLS = ['Inbound','Chat']
N = len(base)
FIN = N + 1                       # última fila de datos en la hoja DATOS

wb = Workbook()

# ══════════════════════════════════════════════════════════════════
#  Utilidades de formato
# ══════════════════════════════════════════════════════════════════
def px(ws, ref, txt, size=11, bold=False, color='1E1235', fill=None,
       align='left', wrap=False, fmt=None, borde=False):
    c = ws[ref]
    c.value = txt
    c.font = Font(name=FUENTE, size=size, bold=bold, color=color)
    if fill: c.fill = PatternFill('solid', fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fmt: c.number_format = fmt
    if borde:
        f = Side(style='thin', color=LN)
        c.border = Border(left=f, right=f, top=f, bottom=f)
    return c

def titulo_hoja(ws, texto, sub, ancho=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    px(ws, 'A1', texto, size=17, bold=True, color='FFFFFF', fill=AZUL)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)
    px(ws, 'A2', sub, size=10, color=TXT)
    ws.row_dimensions[2].height = 20

def encabezados(ws, fila, col0, nombres, anchos=None):
    for i, nm in enumerate(nombres):
        c = px(ws, f'{get_column_letter(col0+i)}{fila}', nm, size=9.5, bold=True,
               color='FFFFFF', fill=AZUL2, align='center', wrap=True, borde=True)
        if anchos: ws.column_dimensions[get_column_letter(col0+i)].width = anchos[i]
    ws.row_dimensions[fila].height = 30

def tarjeta(ws, fila, col, etiqueta, formula, fmt, color=AZUL2, nota=None):
    """Bloque de indicador: rótulo arriba, número grande abajo."""
    a, b = get_column_letter(col), get_column_letter(col+1)
    ws.merge_cells(f'{a}{fila}:{b}{fila}')
    px(ws, f'{a}{fila}', etiqueta, size=9, bold=True, color=TXT,
       fill=LILA_LT, align='center')
    ws.merge_cells(f'{a}{fila+1}:{b}{fila+1}')
    px(ws, f'{a}{fila+1}', formula, size=20, bold=True, color=color,
       align='center', fmt=fmt)
    ws.row_dimensions[fila].height = 18
    ws.row_dimensions[fila+1].height = 30
    if nota:
        ws.merge_cells(f'{a}{fila+2}:{b}{fila+2}')
        px(ws, f'{a}{fila+2}', nota, size=8.5, color=TXT, align='center')

# Rangos con nombre para que las fórmulas se lean solas
R = {c: f"DATOS!${get_column_letter(i+1)}$2:${get_column_letter(i+1)}${FIN}"
     for i, c in enumerate(base.columns)}

def cuenta(cond):
    """COUNTIFS legible a partir de pares (columna, criterio)."""
    partes = []
    for col, crit in cond:
        partes.append(f'{R[col]},{crit}')
    return 'COUNTIFS(' + ','.join(partes) + ')'

def nps_formula(cond):
    """NPS = (promotores - detractores) / total, protegido contra división por cero."""
    tot = cuenta(cond)
    pro = cuenta(cond + [('Clasificacion', '"Promotores"')])
    det = cuenta(cond + [('Clasificacion', '"Detractores"')])
    return f'=IFERROR(({pro}-{det})/{tot},"")'

# ══════════════════════════════════════════════════════════════════
#  HOJA 1 · DATOS  — toda la base en una sola hoja
# ══════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = 'DATOS'
for j, col in enumerate(base.columns, start=1):
    px(ws, f'{get_column_letter(j)}1', col, size=10, bold=True,
       color='FFFFFF', fill=AZUL2, align='center')
for i, fila in enumerate(base.itertuples(index=False), start=2):
    for j, v in enumerate(fila, start=1):
        ws.cell(row=i, column=j, value=v)
anchos = [11, 9, 10, 34, 30, 26, 46, 7, 13]
for j, a in enumerate(anchos, start=1):
    ws.column_dimensions[get_column_letter(j)].width = a
for i in range(2, FIN+1):
    ws.cell(row=i, column=1).number_format = 'dd/mm/yyyy'
tab = Table(displayName='BaseNPS', ref=f'A1:{get_column_letter(len(base.columns))}{FIN}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = 'A2'
ws.sheet_view.showGridLines = False

# ══════════════════════════════════════════════════════════════════
#  HOJA 2 · TABLERO
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet('TABLERO')
ws.sheet_view.showGridLines = False
titulo_hoja(ws, 'Indicador NPS  ·  junio a agosto de 2026',
            'Todas las cifras se calculan sobre la hoja DATOS. Al filtrar o ampliar esa hoja, este tablero se actualiza.', 12)
for j, a in enumerate([16,11,11,11,11,11,11,11,11,11,11,13], start=1):
    ws.column_dimensions[get_column_letter(j)].width = a

# ── Indicadores generales ──
px(ws, 'A4', 'RESULTADO DEL PERÍODO', size=11, bold=True, color=AZUL3)
tarjeta(ws, 5, 1, 'NPS GENERAL', nps_formula([]), '0.0%', AZUL2, 'Promotores − detractores')
tarjeta(ws, 5, 3, 'ENCUESTAS', f'=COUNTA({R["Clasificacion"]})', '#,##0', AZUL2, 'Base del período')
tarjeta(ws, 5, 5, 'DETRACTORES',
        f'=COUNTIF({R["Clasificacion"]},"Detractores")', '#,##0', BAD, 'Calificación baja')
tarjeta(ws, 5, 7, '% DETRACTORES',
        f'=IFERROR(COUNTIF({R["Clasificacion"]},"Detractores")/COUNTA({R["Clasificacion"]}),"")',
        '0.0%', BAD, 'Sobre el total')
tarjeta(ws, 5, 9, 'NPS INBOUND', nps_formula([('Skill','"Inbound"')]), '0.0%', AZUL2, 'Skill de llamadas')
tarjeta(ws, 5, 11, 'NPS CHAT', nps_formula([('Skill','"Chat"')]), '0.0%', BAD, 'Skill de chat')

# ── Comparativa mensual por skill ──
px(ws, 'A10', 'COMPARATIVA MENSUAL POR SKILL', size=11, bold=True, color=AZUL3)
encabezados(ws, 11, 1, ['SKILL','MES','ENCUESTAS','PROMOTORES','NEUTROS',
                        'DETRACTORES','% DETRACTORES','NPS'],
            [16,11,11,11,11,11,13,11])
f = 12
for sk in SKILLS:
    for m in MESES:
        cond = [('Skill', f'"{sk}"'), ('Mes', f'"{m}"')]
        px(ws, f'A{f}', sk, size=10, bold=True, borde=True)
        px(ws, f'B{f}', m, size=10, borde=True)
        px(ws, f'C{f}', f'={cuenta(cond)}', size=10, align='center', fmt='#,##0', borde=True)
        for k, cat in enumerate(['Promotores','Neutros','Detractores']):
            px(ws, f'{get_column_letter(4+k)}{f}',
               '=' + cuenta(cond + [('Clasificacion', '"%s"' % cat)]),
               size=10, align='center', fmt='#,##0', borde=True)
        px(ws, f'G{f}', f'=IFERROR(F{f}/C{f},"")', size=10, align='center', fmt='0.0%', borde=True)
        px(ws, f'H{f}', nps_formula(cond), size=10, bold=True, align='center',
           fmt='0.0%', borde=True)
        f += 1
    # subtotal del skill
    px(ws, f'A{f}', f'Total {sk}', size=10, bold=True, fill=LILA_LT, borde=True)
    px(ws, f'B{f}', '', fill=LILA_LT, borde=True)
    px(ws, f'C{f}', f'=SUM(C{f-3}:C{f-1})', size=10, bold=True, align='center',
       fmt='#,##0', fill=LILA_LT, borde=True)
    for k in range(3):
        L = get_column_letter(4+k)
        px(ws, f'{L}{f}', f'=SUM({L}{f-3}:{L}{f-1})', size=10, bold=True,
           align='center', fmt='#,##0', fill=LILA_LT, borde=True)
    px(ws, f'G{f}', f'=IFERROR(F{f}/C{f},"")', size=10, bold=True, align='center',
       fmt='0.0%', fill=LILA_LT, borde=True)
    px(ws, f'H{f}', nps_formula([('Skill', f'"{sk}"')]), size=10, bold=True,
       align='center', fmt='0.0%', fill=LILA_LT, borde=True)
    f += 1

# ── Gráfico de evolución mensual ──
ch = LineChart()
ch.title = 'NPS mensual por skill'
ch.y_axis.numFmt = '0%'
ch.y_axis.title = 'NPS'
ch.height, ch.width = 7.6, 15
datos_in = Reference(ws, min_col=8, min_row=12, max_row=14)   # Inbound
datos_ch = Reference(ws, min_col=8, min_row=16, max_row=18)   # Chat
cats     = Reference(ws, min_col=2, min_row=12, max_row=14)
ch.add_data(datos_in, titles_from_data=False)
ch.add_data(datos_ch, titles_from_data=False)
ch.set_categories(cats)
ch.series[0].tx = None; ch.series[1].tx = None
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
ch.series[0].tx = SeriesLabel(v='Inbound')
ch.series[1].tx = SeriesLabel(v='Chat')
ch.series[0].graphicalProperties.line.width = 28000
ch.series[1].graphicalProperties.line.width = 28000
ch.series[0].graphicalProperties.line.solidFill = AZUL2
ch.series[1].graphicalProperties.line.solidFill = BAD
ch.series[0].smooth = False; ch.series[1].smooth = False
ws.add_chart(ch, 'J10')

px(ws, 'A21', 'Nota: Inbound y Chat usan escalas de calificación distintas '
   '(Inbound 0 a 9, Chat 0 a 10), por lo que no son comparables entre sí. '
   'Cada uno debe leerse contra su propia historia.', size=9.5, color=TXT)
ws.merge_cells('A21:H21')

# ══════════════════════════════════════════════════════════════════
#  HOJAS 3-4 · TIPIFICACIÓN POR SKILL
# ══════════════════════════════════════════════════════════════════
for sk in SKILLS:
    ws = wb.create_sheet(f'TIPIFICACIÓN {sk.upper()}')
    ws.sheet_view.showGridLines = False
    titulo_hoja(ws, f'Tipificación · {sk}',
        'Detractores aportados por cada proceso y subproceso, con el resultado de cada mes. '
        'Tasa = detractores sobre encuestas de esa misma tipificación.', 11)
    sub = base[base['Skill'] == sk]
    fila_ini = 5
    for nivel, col_dato, tope, minimo in [('Proceso','Proceso',12,0),
                                          ('Subproceso','Subproceso',12,0)]:
        det = sub[sub['Clasificacion'] == 'Detractores']
        orden = det[col_dato].value_counts().head(tope).index.tolist()
        px(ws, f'A{fila_ini-1}', f'POR {nivel.upper()}  ·  ordenado por detractores aportados',
           size=11, bold=True, color=AZUL3)
        encabezados(ws, fila_ini, 1,
            [nivel.upper(),'ENCUESTAS','DETRACTORES','% DEL TOTAL','TASA','NPS',
             'DET. JUNIO','DET. JULIO','DET. AGOSTO','NPS JUNIO','NPS AGOSTO'],
            [44,11,12,11,10,10,11,11,12,11,12])
        f = fila_ini + 1
        for nombre in orden:
            cond = [('Skill', f'"{sk}"'), (col_dato, f'"{nombre}"')]
            px(ws, f'A{f}', nombre, size=9.5, borde=True, wrap=True)
            px(ws, f'B{f}', '=' + cuenta(cond), size=10, align='center', fmt='#,##0', borde=True)
            px(ws, f'C{f}', '=' + cuenta(cond + [('Clasificacion','"Detractores"')]),
               size=10, align='center', fmt='#,##0', borde=True)
            px(ws, f'D{f}',
               f'=IFERROR(C{f}/COUNTIFS({R["Skill"]},"{sk}",{R["Clasificacion"]},"Detractores"),"")',
               size=10, align='center', fmt='0.0%', borde=True)
            px(ws, f'E{f}', f'=IFERROR(C{f}/B{f},"")', size=10, align='center',
               fmt='0.0%', borde=True)
            px(ws, f'F{f}', nps_formula(cond), size=10, bold=True, align='center',
               fmt='0.0%', borde=True)
            for k, m in enumerate(MESES):
                px(ws, f'{get_column_letter(7+k)}{f}',
                   '=' + cuenta(cond + [('Mes', f'"{m}"'), ('Clasificacion','"Detractores"')]),
                   size=10, align='center', fmt='#,##0', borde=True)
            px(ws, f'J{f}', nps_formula(cond + [('Mes','"Junio"')]),
               size=10, align='center', fmt='0.0%', borde=True)
            px(ws, f'K{f}', nps_formula(cond + [('Mes','"Agosto"')]),
               size=10, align='center', fmt='0.0%', borde=True)
            f += 1
        # Semáforo sobre el NPS y barras sobre los detractores
        ws.conditional_formatting.add(f'F{fila_ini+1}:F{f-1}',
            CellIsRule(operator='lessThan', formula=['0.5'],
                       font=Font(name=FUENTE, size=10, bold=True, color=BAD)))
        ws.conditional_formatting.add(f'C{fila_ini+1}:C{f-1}',
            DataBarRule(start_type='min', end_type='max', color=LILA))
        fila_ini = f + 3
    ws.freeze_panes = 'B6'

# ══════════════════════════════════════════════════════════════════
#  HOJAS 5-6 · AGENTES POR SKILL
# ══════════════════════════════════════════════════════════════════
MIN_ENC = 50
for sk in SKILLS:
    ws = wb.create_sheet(f'AGENTES {sk.upper()}')
    ws.sheet_view.showGridLines = False
    titulo_hoja(ws, f'Agentes · {sk}',
        f'Agentes con {MIN_ENC} o más encuestas en el período, ordenados por NPS. '
        'Cada uno se compara contra el NPS de su propio skill, no contra el total.', 11)
    sub = base[(base['Skill'] == sk) & (base['Agente'] != 'Sin asignar')]
    cnt = sub['Agente'].value_counts()
    elegibles = cnt[cnt >= MIN_ENC].index.tolist()
    # se ordenan aquí, porque una hoja no puede ordenarse sola
    def npsx(a):
        s = sub[sub['Agente'] == a]
        return (s['Clasificacion'].eq('Promotores').sum()
                - s['Clasificacion'].eq('Detractores').sum()) / len(s)
    elegibles.sort(key=npsx)

    px(ws, 'A4', f'NPS DEL SKILL {sk.upper()}', size=10, bold=True, color=TXT, fill=LILA_LT)
    px(ws, 'B4', nps_formula([('Skill', f'"{sk}"')]), size=13, bold=True,
       color=AZUL2, align='center', fmt='0.0%')
    ws['B4'].comment = Comment(
        'Referencia del skill. La columna Diferencia mide cada agente contra este valor.',
        'Análisis NPS')

    encabezados(ws, 6, 1,
        ['#','AGENTE','ENCUESTAS','PROMOTORES','DETRACTORES','% DETRACTORES',
         'NPS','DIFERENCIA','NPS JUNIO','NPS JULIO','NPS AGOSTO'],
        [5,40,11,12,12,13,10,11,11,11,12])
    f = 7
    for i, a in enumerate(elegibles, start=1):
        cond = [('Skill', f'"{sk}"'), ('Agente', f'"{a}"')]
        px(ws, f'A{f}', i, size=10, align='center', borde=True)
        px(ws, f'B{f}', a, size=9.5, borde=True)
        px(ws, f'C{f}', '=' + cuenta(cond), size=10, align='center', fmt='#,##0', borde=True)
        px(ws, f'D{f}', '=' + cuenta(cond + [('Clasificacion','"Promotores"')]),
           size=10, align='center', fmt='#,##0', borde=True)
        px(ws, f'E{f}', '=' + cuenta(cond + [('Clasificacion','"Detractores"')]),
           size=10, align='center', fmt='#,##0', borde=True)
        px(ws, f'F{f}', f'=IFERROR(E{f}/C{f},"")', size=10, align='center',
           fmt='0.0%', borde=True)
        px(ws, f'G{f}', nps_formula(cond), size=10, bold=True, align='center',
           fmt='0.0%', borde=True)
        px(ws, f'H{f}', f'=IFERROR(G{f}-$B$4,"")', size=10, bold=True, align='center',
           fmt='+0.0%;-0.0%', borde=True)
        for k, m in enumerate(MESES):
            px(ws, f'{get_column_letter(9+k)}{f}', nps_formula(cond + [('Mes', f'"{m}"')]),
               size=10, align='center', fmt='0.0%', borde=True)
        f += 1
    ws.conditional_formatting.add(f'H7:H{f-1}',
        CellIsRule(operator='lessThan', formula=['-0.1'],
                   font=Font(name=FUENTE, size=10, bold=True, color=BAD),
                   fill=PatternFill('solid', fgColor='FEE2E2')))
    ws.conditional_formatting.add(f'G7:G{f-1}',
        DataBarRule(start_type='min', end_type='max', color=LILA))
    ws.freeze_panes = 'C7'
    px(ws, f'A{f+1}',
       f'Se excluyen los agentes con menos de {MIN_ENC} encuestas en el período: con una base '
       'pequeña, un puñado de respuestas determina la posición. El campo Supervisor no se usa '
       'porque figura sin asignar en la totalidad de los registros.',
       size=9, color=TXT)

# ══════════════════════════════════════════════════════════════════
#  HOJA 7 · METODOLOGÍA
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet('METODOLOGÍA')
ws.sheet_view.showGridLines = False
titulo_hoja(ws, 'Metodología y alcance',
            'Origen de las cifras y diferencia frente a las hojas de resumen de los archivos originales.', 8)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 110
bloques = [
 ('Fuente de los datos',
  'Hoja «BBDD Encuestas» de los consolidados mensuales de junio, julio y agosto de 2026. '
  'La hoja DATOS de este archivo contiene los 40.738 registros individuales, sin agregar.'),
 ('Por qué difieren las otras hojas del archivo original',
  'Las hojas de resumen (Detallado NPS Agente, Detallado NPS Proceso, Profundización) provienen de una '
  'tabla dinámica cuya caché conserva 123.031 registros de 2025 y 2026. La caché apunta a «BBDD Encuestas» '
  'pero quedó sin actualizar, de modo que muestra un acumulado histórico y no el mes del archivo. '
  'Por eso sus cifras no coinciden con las de este informe.'),
 ('Cálculo del indicador',
  'NPS = porcentaje de promotores menos porcentaje de detractores, sobre las encuestas con clasificación. '
  'Los neutros integran la base pero no suman ni restan. Todas las celdas de los tableros son fórmulas '
  'sobre la hoja DATOS: si se amplía o corrige esa hoja, los tableros se recalculan.'),
 ('Escalas de calificación',
  'Inbound califica de 0 a 9 (promotor 8-9, neutro 6-7, detractor 0-5). Chat califica de 0 a 10 '
  '(promotor 9-10, neutro 7-8, detractor 0-6). Una misma nota no equivale a la misma categoría en los dos '
  'skills, por lo que se utiliza la clasificación registrada y no la nota cruda, y los skills no se comparan entre sí.'),
 ('Mínimos aplicados',
  f'Agentes: {MIN_ENC} encuestas en el período. Los procesos y subprocesos se muestran completos, '
  'ordenados por detractores aportados; al leer la columna NPS conviene revisar antes la de encuestas.'),
 ('Definición de las columnas',
  '% del total = participación de esa tipificación en los detractores del skill (dónde está el volumen). '
  'Tasa = detractores sobre encuestas de esa misma tipificación (qué tan mal califican dentro de ella). '
  'Diferencia = NPS del agente menos el NPS de su skill.'),
]
f = 4
for t, c in bloques:
    px(ws, f'A{f}', t, size=10.5, bold=True, color=AZUL2, fill=LILA_LT, wrap=True)
    px(ws, f'B{f}', c, size=10, color='1E1235', wrap=True)
    ws.row_dimensions[f].height = 46
    f += 1
px(ws, f'A{f+1}', 'Elaborado por', size=10, bold=True, color=TXT)
px(ws, f'B{f+1}', 'Supervisión Skill Inbound · agosto de 2026', size=10, color=TXT)

wb.active = 1
wb.save('xls/Tablero_NPS_Junio_Agosto_2026.xlsx')
print('generado · hojas:', wb.sheetnames)
